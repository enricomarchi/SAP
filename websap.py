from selenium import webdriver
from selenium.webdriver import chrome
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.support.wait import WebDriverWait
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import InvalidSelectorException
import pandas as pd
import time
import openpyxl
import numpy as np

class Ex_VDT_non_trovata(Exception):
    def __init__(self, vdt, messaggio):
        super().__init__(messaggio)
        self.vdt = vdt

class WebSAP:
    DEF_tempo_operazione: float = 1.5
    DEF_timeout: float = 600

    def __init__(
            self,
            utente: str,
            password: str
    ):
        """
        Lancia l'applicativo Web PS2 e si connette con il nome utente e password forniti.
        :param utente: nome utente
        :param password: password
        """
        self.utente = utente
        self.password = password
        self.driver = None
        self.tempo_operazione = WebSAP.DEF_tempo_operazione
        self.timeout = WebSAP.DEF_timeout
        self.wb = None
        self.ws = None
        self.df_tariffe = pd.read_parquet("Tariffe.parquet")
        self.df_macep = pd.read_parquet("Macep.parquet")
        self.logon()

    def logon(
            self
    ):
        """
        Esegue il logon con l'utenza e la password fornita durante la creazione dell'oggetto
        :return:
        """
        self.driver: chrome = webdriver.Chrome()
        self.driver.maximize_window()
        self.driver.get("https://ps2.rfi.it/sap/bc/webdynpro/sap/z0029_wd_rfi_main_comp?sap-client=111&sap-language=IT#")
        self.testo(self.utente, '//input[@id="sap-user"]')  # Utente
        self.testo(self.password, '//input[@id="sap-password"]')  # Password
        self.click('//div[@id="LOGON_BUTTON"]')  # LOGON

        lista = self.driver.find_elements(By.XPATH, '//span[contains(text(), "utente è già collegato al sistema con le seguenti sessioni")]')
        if len(lista) > 0:
            self.click('//span[text()="Cont."]/../..')

        self.click('//div[@id="SYSTEM_MESSAGE_CONTINUE_BUTTON"]')

    def elemento(
            self,
            xpath: str
    ):
        """
        Ritorna l'elemento HTML con xpath indicato.
        :param xpath: xpath dell'elemento desiderato
        :return: WebElement
        """
        for i in range(1, int(self.timeout)):
            try:
                elem = self.driver.find_element(By.XPATH, xpath)
                time.sleep(1)
                return elem
            except NoSuchElementException:
                pass

        #elem = WebDriverWait(self.driver, self.timeout, 1).until(lambda x: x.find_element(By.XPATH, xpath))

    def lista_elementi(
            self,
            xpath: str
    ):
        """
        Ritorna una lista di elementi HTML con xpath indicato.
        :param xpath: xpath degli elementi desiderati
        :return: WebElement
        """
        lista = []
        for i in range(1, int(self.timeout)):
            try:
                lista = self.driver.find_elements(By.XPATH, xpath)
                return lista
            except NoSuchElementException:
                print(f"NoSuchElementException, xpath:{xpath}, def lista_elementi, riga 89")
                time.sleep(1)

        return lista

    def attesa_caricamento(
            self,
    ):
        """
        Attende il caricamento della pagina, fino a quando non scompare la ruota blu che gira.
        """
        time.sleep(1)
        elem: WebElement = self.elemento('//div[@id="ur-loading"]')
        while elem.value_of_css_property("visibility") == "visible":
            time.sleep(1)

    def click(
            self,
            xpath: str,
    ):
        """
        Simula un click del mouse sull'elemento puntato da xpath.
        :param xpath: xpath dell'elemento da cliccare
        :return: WebElement
        """
        elem = self.elemento(xpath)
        elem.click()
        time.sleep(1)
        return elem

    def testo(
            self,
            txt: str,
            xpath: str
    ):
        """
        Fa un click sull'elemento ed inserisce il testo specificato.
        :param txt: il testo da inserire
        :param xpath: xpath dell'elemento
        :return: WebElement
        """
        elem = self.click(xpath)
        elem.clear()
        time.sleep(1)
        elem.send_keys(txt)
        time.sleep(1)
        return elem
                    

    def leggi_excel(self, file_excel: str, nome_foglio: str = ""):
        self.wb = openpyxl.load_workbook(file_excel)
        self.ws = self.wb.active if nome_foglio == "" else self.wb[nome_foglio]      

        # Legge l'header nella riga 1
        colonna = 1
        finecolonne = False
        colonne = []
        while not finecolonne:
            cella = self.ws.cell(row=1, column=colonna)
            finecolonne = (cella.value == None or cella.value == "")
            if not finecolonne:
                colonne.append(cella.value)
            colonna += 1

        # Legge i dati
        df = pd.DataFrame(columns=colonne)
        for riga in range(2, self.ws.max_row + 1):
            for colonna in range(len(colonne)):
                df.loc[riga, colonne[colonna]] = self.ws.cell(row=riga, column=colonna+1).value
        return df    

    def sal(self, oda: str, file_excel: str):
        # Entra in WebSAL
        time.sleep(1)
        self.click('//div[@title="SAL"]')           # Tasto SAL
        time.sleep(1)
        self.click('//td[@ut="3"][@cc="3"]/a')      # Tasto freccia Assistente
        time.sleep(1)
        self.click('//span[text()="OK"]/../..')     # OK al popup
        self.click('//div[text()="GESTIONE"]')      # Scheda GESTIONE
        self.attesa_caricamento()
        self.testo(oda, '//span[text()="NUMERO ORDINE DI ACQUISTO NOTO"]/../../../following-sibling::td//input')  # ODA

        df = self.leggi_excel(file_excel=file_excel, nome_foglio="VDT") #pd.read_excel(io=file_excel, sheet_name="VDT", header=0)
        df = df[(df["Inserita"] == "") | (pd.isna(df["Inserita"]))]
        df = df.dropna(axis=0, subset=['Quantità'])
        df.VDT = df.VDT.str.upper()
        df.VDT = df.VDT.fillna("")
        df.NV = df.NV.fillna("")
        df.Descrizione = df.Descrizione.fillna("")
        df.Inserita = df.Inserita.fillna("")
        posizioni = df.loc[df["Inserita"] == "", "Posizione"].unique()
        tot_vdt = len(df.loc[df["Inserita"] == "", "VDT"])
        i_vdt = 1
        for pos in posizioni:
            self.testo(str(pos), '//span[text()="Posizione Numero"]/../../../following-sibling::td//input')  # Posizione
            df_posizione = df.loc[df["Posizione"] == pos, :]
            parti_opera = df_posizione.loc[df_posizione["Inserita"] == "", "PO"].unique()
            for po in parti_opera:
                self.testo(str(po), '//span[contains(text(), "Opera")]/../../../following-sibling::td//input')   # Parte d'opera
                df_parte_opera = df_posizione.loc[df_posizione["PO"] == po, :]
                self.click('//div[@title="Vai alla Gestione delle Voci di Tariffa"]')
                self.attesa_caricamento()

                # Se compare una richiesta di modificare la versione in elaborazione clicca e prosegui
                time.sleep(1)
                lista = self.lista_elementi('//span[text()="Ultima versione IN ELABORAZIONE"]/..')
                if len(lista) > 0:
                    lista[0].click()  # Checkbox "ultima versione in elaborazione"
                    self.click('//span[text()="OK"]/../..')  # Tasto OK
                    self.attesa_caricamento()

                time.sleep(1)
                for index, row in df_parte_opera.iterrows():
                    if row.Inserita == "":
                        try:
                            print(f"\rVDT n.{i_vdt} di {tot_vdt}", end=" ", flush=True)
                            self.__inserisci_vdt_sal(index, row)
                            time.sleep(1)
                            df.at[index, "Inserita"] = "x"
                            self.ws.cell(row=index, column=df.columns.get_loc("Inserita")+1).value = "x"
                        except InvalidSelectorException as e:
                            print(str(e))
                            input('Premere INVIO per proseguire')
                        except Ex_VDT_non_trovata as e:
                            df.at[index, "Inserita"] = "manca"
                            self.ws.cell(row=index, column=df.columns.get_loc("Inserita")+1).value = "manca"
                            print(str(e))
                        finally:
                            self.wb.save(file_excel)
                            i_vdt += 1

                self.click('//span[text()="Salva"]/../..')              # SALVA
                self.click('//span[text()="Altra gestione"]/../..')     # Altra gestione                
                self.wb.close()

    def __inserisci_vdt_sal(self, index, row):
        vdt = row.VDT
        q = row.Quantità
        data = row.Data
        ribasso = row.Rib
        descrizione = row.Descrizione
        nuova_voce = row.NV
        ed_tariffa = row.Ed_Tariffa

        print(vdt)
        lista = self.lista_elementi(f'//span[text()="{vdt}"]')  # Verifica se la VDT è già stata inserita
        descrizione_vdt = vdt
        time.sleep(1)
        if len(lista) == 0:                                 # Se non è stata già inserita la inserisce
            self.click('//div[@title="Seleziona una nuova Voce di Tariffa"]')
            time.sleep(1)
            self.click('//div[text()="Contr.Catalogo"]')
            time.sleep(1)
            if nuova_voce == "":
                self.testo(vdt, '//div[@class="urPWContent"]//span[text()="VdT"]/../../../../td[2]//input')  # campo VDT
                elem = self.elemento('//div[@class="urPWContent"]//span[text()="Descrizione"]/../../../../td[4]//input')  # cancella il campo descrizione
                elem.clear()
            else:
                self.testo(f'*{vdt}*', '//div[@class="urPWContent"]//span[text()="Descrizione"]/../../../../td[4]//input')  # campo Descrizione
                elem = self.elemento('//div[@class="urPWContent"]//span[text()="VdT"]/../../../../td[2]//input')  # cancella il campo VDT
                elem.clear()
            time.sleep(1)
            self.click('//span[text()="Visualizza"]/../..')     # visualizza

    # ============================================ SELEZIONE EDIZIONE TARIFFA =========================================
            try:
                time.sleep(1)
                versione_vdt = self.trova_versione_vdt(vdt, ed_tariffa, nuova_voce)
            except Ex_VDT_non_trovata:
                time.sleep(1)
                self.click('//span[text()="Chiudi"]/../..')
                raise
            time.sleep(1)
            versioni = self.lista_elementi(
                f'//span[text()="Visualizza"]/../../../../../../../../../../../tr[2]//td//td//tr[2]//tr[@rr="{versione_vdt}"]/td[@cc="0"]'
            )
            time.sleep(1)
            if len(versioni) > 0:
                versioni[0].click()   # checkbox voce
                elem = self.elemento(
                    f'//span[text()="Visualizza"]/../../../../../../../../../../../tr[2]//td//td//tr[2]//tr[@rr="{versione_vdt}"]/td[3]/span/span')
            else:  # in caso contrario preme la prima della lista, che dovrebbe essere l'unica dell'elenco
                self.click(f'//span[text()="Visualizza"]/../../../../../../../../../../../tr[2]//td//td//tr[2]//tr[@rr="1"]/td')
                elem = self.elemento(f'//span[text()="Visualizza"]/../../../../../../../../../../../tr[2]//td//td//tr[2]//tr[@rr="1"]/td[3]/span/span')
            if nuova_voce != "":
                descrizione_vdt = elem.text
            time.sleep(1)
    # =================================================================================================================
    
            self.click('//span[text()="Inserisci"]/../..')      # inserisci

        time.sleep(1)
        self.click(f'//span[text()="{descrizione_vdt}"]/../../../../../../../../../../../../td[2]//span[text()="Gestione Misurazioni"]/../..')  # Gestione misurazioni

        time.sleep(1)
        tabella_misure = '//span[text()="+/-"]/../../../../..'  # xpath della tabella misure
        tot_righe = self.lista_elementi(f'{tabella_misure}/tr')  # N.B. la riga n. 1 è l'header
        # Trova le colonne interessate nell'header
        time.sleep(1)
        header = self.lista_elementi('//span[text()="+/-"]/../../../../../tr[1]/th')
        indice_link_descrizione = 0
        indice_misura = 0
        indice_data = 0
        time.sleep(1)
        for i in range(1, len(header)):
            elem = self.elemento(f'//span[text()="+/-"]/../../../../../tr[1]/th[{i}]//span/span')
            time.sleep(1)
            if elem.text == "Nota":
                indice_link_descrizione = i
            elif elem.text == "Totale":
                indice_misura = i
            elif elem.text[0:4] == "Data":
                indice_data = i
                break
        xpath_data = ""
        xpath_link_descrizione = ""
        xpath_misura = ""

        # Cerca la prima riga vuota dove inserire la misura
        time.sleep(1)
        for i in range(2, len(tot_righe)):   # la riga n.1 è l'header
            #time.sleep(1)
            xpath_misura = f'{tabella_misure}/tr[{i}]/td[{indice_misura}]//input'
            campo_misura: WebElement = self.elemento(xpath_misura)
            xpath_data = f'{tabella_misure}/tr[{i}]/td[{indice_data}]//input'
            xpath_link_descrizione = f'{tabella_misure}/tr[{i}]/td[{indice_link_descrizione}]//a'
            if campo_misura.get_attribute('value') == "":   # quando trova la riga vuota esce dal ciclo
                break

        self.testo(data, xpath_data)
        time.sleep(1)
        self.testo(str(q) + Keys.ENTER, xpath_misura)
        time.sleep(1)
        if descrizione != "":
            self.click(xpath_link_descrizione)
            self.testo(descrizione, '//textarea')
            self.click('//span[text()="Salva"]/../..')  # Salva
        time.sleep(1)
        self.click('//span[text()="Torna alla SAL"]/../..')     # Torna alla sal
        if ribasso == "NO":
            time.sleep(2)
            self.click(f'//span[text()="{vdt}"]/../../../../../../../../../../../../..//span[text()="No"]/..')
            time.sleep(2)
            self.click('//span[text() = "Aggiorna"] /../..')
            self.attesa_caricamento()
            time.sleep(2)

    def trova_versione_vdt(self, vdt, ed_tariffa, nuova_voce):
        tot_righe_elem = self.lista_elementi('//span[text()="Visualizza"]/../../../../../../../../../../../tr[2]//td//td//tr[2]//tr/td/table/tbody/tr')
        tot_righe = len(tot_righe_elem) - 1 # -1 perché una riga è l'header 
        if tot_righe == 0:
            raise Ex_VDT_non_trovata(vdt, f"La VDT {vdt} edizione {ed_tariffa} non è presente in web sal.")
        if nuova_voce == "MaCeP":
            vdt = vdt.replace("*", "")
            prezzo_da_trovare = self.df_macep.query("Codice_Materiale == @vdt")[ed_tariffa].iloc[0]
        elif nuova_voce == "x":
            return 1
        else:
            prezzo_da_trovare = self.df_tariffe.query("Numero_S_VdT == @vdt")[ed_tariffa].iloc[0]
        trovata = False
        for riga in range(1, tot_righe + 1):
            time.sleep(1)
            pr_un = self.elemento(f'//span[text()="Visualizza"]/../../../../../../../../../../../tr[2]//td//td//tr[2]//tr[@rr="{riga}"]/td[@cc="6"]')
            time.sleep(1)
            pr_un = pr_un.text
            pr_un = pr_un.replace(".", "").replace(",", ".")
            pr_un = float(pr_un)
            if pr_un == prezzo_da_trovare:
                trovata = True
                break
        if trovata == False:
            raise Ex_VDT_non_trovata(vdt, f"La VDT {vdt} edizione {ed_tariffa} non è presente in web sal.")
        
        return riga
            
    def perizia(self, network: str, file_excel: str, nome_foglio: str):
        # Entra in Web Perizie
        time.sleep(1)
        self.click('//div[@title="Perizie"]')       # Tasto perizie
        time.sleep(1)
        self.click('//td[@ut="3"][@cc="3"]/a')      # Tasto freccia Specialista
        self.attesa_caricamento()
        self.click('//span[text()="OK"]/../..')     # OK al popup
        self.click('//div[text()="GESTIONE VPR"]')      # Scheda GESTIONE
        self.attesa_caricamento()
        self.testo(network, '//span[text()="NUMERO DI NETWORK"]/../../../following-sibling::td//input')  # Network

        df = self.leggi_excel(file_excel=file_excel, nome_foglio=nome_foglio) #df = pd.read_excel(io=file_excel, sheet_name="VDT", header=0)
        df = df[(df["Inserita"] == "") | (pd.isna(df["Inserita"]))]
        df.VDT = df.VDT.str.upper()
        df.VDT = df.VDT.fillna("")
        df.Descrizione = df.Descrizione.fillna("")
        df = df[df["Quantità"] != 0]
        df.Inserita = df.Inserita.fillna("")
        df.Prezzo = df.Prezzo.fillna(0)
        operazioni = df.Operazione.unique()
        for op in operazioni:
            self.testo(str(op), '//span[text()="OPERAZIONE"]/../../../following-sibling::td//input')  # Operazione
            df_operazione = df[df["Operazione"] == op]
            parti_opera = df_operazione.PO.unique()
            for po in parti_opera:
                elem = self.testo(str(po), '//span[starts-with(text(), "PARTE")]/../../../following-sibling::td//input')   # Parte d'opera
                df_parte_opera = df_operazione[df_operazione["PO"] == po]
                elem.send_keys(Keys.RETURN)
                self.click('//span[text()="GESTIONE RISORSE"]/../..')
                self.attesa_caricamento()

                # Se compare un messaggio che un altro utente sta modificando questa VPR, seleziona di elaborare questa VPR
                time.sleep(1)
                lista = self.lista_elementi('//span[starts-with(text(), "Continuare con questa V.P.R.")]/..')
                if len(lista) > 0:
                    lista[0].click()  # Checkbox "Continuare con questa VPR"
                    self.click('//span[text()="PROCEDI"]/../..')  # Tasto PROCEDI
                    self.attesa_caricamento()

                # Se compare un messaggio che sono state rilevate versioni non salvate
                time.sleep(1)
                lista = self.lista_elementi('//span[starts-with(text(), "Ultima versione IN ELABORAZIONE")]')
                if len(lista) > 0:
                    lista[0].click()  # Checkbox "Ultima versione in elaborazione"
                    self.click('//span[text()="PROCEDI"]/../..')  # Tasto PROCEDI
                    self.attesa_caricamento()

                time.sleep(1)
                for index, row in df_parte_opera.iterrows():
                    if row.Inserita == "":
                        try:
                            self.__inserisci_vdt_perizia(index, row)
                            df.at[index, "Inserita"] = "x"
                            self.ws.cell(row=index, column=df.columns.get_loc("Inserita")+1).value = "x"
                        except InvalidSelectorException as e:
                            pass
                        finally:
                            self.wb.save(file_excel)
                            
                self.click('//span[text()="SALVA"]/../..')              # SALVA
                self.click('//span[text()="ALTRA GESTIONE"]/../..')     # Altra gestione
                self.wb.close()
                
    def __inserisci_vdt_perizia(self, index, row):
        vdt = row.VDT
        q = row.Quantità
        prezzo=row.Prezzo
        vdt_standard = row.TipoVDT == "ANAGRAFICA"
        descrizione = row.Descrizione
        UM = row.UM
        
        # Verifica se la VDT è già stata inserita
        print(vdt)
        time.sleep(1)
        if vdt_standard:
            lista = self.lista_elementi(f'//input[@value="{vdt}"]')
        else:
            lista = self.lista_elementi(f'//span[text()="{vdt}"]')

        time.sleep(1)
        if len(lista) == 0:  # Se non è stata già inserita la inserisce
            self.click('//span[text()="[SEL. VdT]"]/..')  # Sel. VDT
            time.sleep(1)
            if vdt_standard:
                self.testo(vdt, '//div[@class="urPWContent"]//span[text()="Voce di Tariffa"]/../../../td[5]//input')  # campo VDT
                self.click('//span[text()="VISUALIZZA"]/../..')  # visualizza
                time.sleep(1)
                self.click(f'//span[text()="{vdt}"]/../../../td')  # checkbox prima voce
            else:
                self.click('//div[text()="NON CODIFICATE"]')
                self.testo(vdt, '//div[@class="urPWContent"]//span[text()="DESCRIZIONE"]/../../../../../../../../../../tr[2]/td[3]//input')  # campo Descrizione
                self.testo(UM, '//div[@class="urPWContent"]//span[text()="DESCRIZIONE"]/../../../../../../../../../../tr[2]/td[4]//input')  # campo UM
            time.sleep(1)
            self.click('//span[text()="INSERISCI"]/../..')  # inserisci
        time.sleep(1)
        if vdt_standard:
            self.click(f'//input[@value="{vdt}"]/../../../td[8]/a')  # Gestione misurazioni
        else:
            # se è una VDT non codificata inserisce anche l'importo unitario
            self.testo(str(prezzo), f'//span[text()="{vdt}"]/../../../../../../../../../../../../../../../../../../tr[2]//span[text()="Prezzo Lordo :"]/../../../td[4]//input')   # campo Prezzo
            time.sleep(1)
            self.click(f'//span[text()="{vdt}"]/../../../td[8]/a')  # Gestione misurazioni

        time.sleep(1)
        tabella_misure = '//span[text()="+/-"]/../../../../..'  # xpath della tabella misure
        tot_righe = []
        try:
            tot_righe = self.lista_elementi(f'{tabella_misure}/tr')  # N.B. la riga n. 1 è l'header
        except TimeoutException as e:
            print(f"TimeoutException, xpath:{tabella_misure}/tr , def __inserisci_vdt_perizia, riga 391")
        xpath_link_descrizione = ""
        xpath_misura = ""
        # Cerca la prima riga vuota dove inserire la misura
        time.sleep(1)
        for i in range(2, len(tot_righe)):  # la riga n.1 è l'header
            xpath_misura = f'{tabella_misure}/tr[{i}]/td[5]//input'
            campo_misura: WebElement = self.elemento(xpath_misura)
            xpath_link_descrizione = f'{tabella_misure}/tr[{i}]/td[3]/div'
            if campo_misura.get_attribute('value') == "":  # quando trova la riga vuota esce dal ciclo
                break

        time.sleep(1)
        self.testo(str(q) + Keys.ENTER, xpath_misura)
        if descrizione != "":
            self.click(xpath_link_descrizione)
            self.testo(descrizione, '//textarea')
            self.click('//span[text()="[SALVA]"]/../..')  # Salva
        time.sleep(1)
        self.click('//span[text()="[TORNA ALLA VPR]"]/../..')  # Torna alla VPR
